import io
import pyqrcode
import eel
from openpyxl import *
from datetime import datetime

eel.init('web')
# opening the existing excel file 
wb = load_workbook('C:\\Users\\ofkoo_000\\Desktop\\Nail-Consent\\consent2.xlsx') 

# create the sheet object 
sheet = wb.active 



def excel(): 
	
	# resize the width of columns in 
	# excel spreadsheet 
    sheet.column_dimensions['A'].width = 30
    sheet.column_dimensions['B'].width = 30
    sheet.column_dimensions['C'].width = 30
    sheet.column_dimensions['D'].width = 20
    sheet.column_dimensions['E'].width = 20
    sheet.column_dimensions['F'].width = 40
    sheet.column_dimensions['G'].width = 50

	# write given data to an excel spreadsheet 
	# at particular location 
    sheet.cell(row=1, column=1).value = "Name"
    sheet.cell(row=1, column=2).value = "LName"
    sheet.cell(row=1, column=3).value = "Phone"
    sheet.cell(row=1, column=4).value = "Signature"
    sheet.cell(row=1, column=5).value = "Date"
    sheet.cell(row=1, column=6).value = "Agree"
    sheet.cell(row=1, column=7).value = "Email id"
    sheet.cell(row=1, column=8).value = "Address"

@eel.expose
def dummy(dummy_param):
    print("I got a parameter: ", dummy_param)
    return "string_value", 1, 1.2, True, [1, 2, 3, 4], {"name": "eel"}


@eel.expose
def generate_qr(fullname, phone, sign,checkbox1, lname):
# assigning the max row and max column 
    # value upto which data is written 
    # in an excel sheet to the variable 
    current_row = sheet.max_row 
    current_column = sheet.max_column 

    # get method returns current text 
    # as string which we write into 
    # excel spreadsheet at particular location 
    sheet.cell(row=current_row + 1, column=1).value = fullname
    sheet.cell(row=current_row + 1, column=2).value = lname
    sheet.cell(row=current_row + 1, column=3).value = phone
    sheet.cell(row=current_row + 1, column=4).value = sign
    sheet.cell(row=current_row + 1, column=5).value = datetime.date(datetime.now())
    sheet.cell(row=current_row + 1, column=6).value = checkbox1


    # save the file 
    wb.save('C:\\Users\\ofkoo_000\\Desktop\\Nail-Consent\\consent2.xlsx') 

   

eel.start('index.html', size=(1000, 600))
